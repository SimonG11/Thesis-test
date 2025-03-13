from openpyxl import load_workbook
import json


def extract_values(s: str):
    # Split the string into two parts at the first comma
    parts = [part.strip() for part in s.split(",", -1)]
    
    # Remove "%" from the first part and convert to integer
    first_value = parts[0]
    if first_value.lower() == "no":
        dep = []
    else:
        dep = first_value.replace("%", "").split(",")
        for i in range(len(dep)):
            try:
                dep[i] = int(dep[i])
            except:
                dep[i] = dep[i] + "- ERROR"
    
    
    # For the second part, split by space and take the second element (e.g., "max 6" -> ["max", "6"])
    try:
        max = int(parts[1].split()[1])
    except:
        try:
            max = int(parts[1].split()[0])
        except:
            max = "ERROR"
    
    return dep, max


def process_excel_sheet(file_path: str, sheet_name: str) -> None:
    """
    Open an Excel file, select a specific worksheet, and iterate over each row.

    Parameters:
        file_path (str): Path to the Excel (.xlsx) file.
        sheet_name (str): The name of the worksheet to process.
    """
    # Load the workbook
    workbook = load_workbook(filename=file_path, data_only=True)
    
    # Select the desired worksheet by name
    if sheet_name not in workbook.sheetnames:
        print(f"Sheet '{sheet_name}' not found in workbook.")
        return
    sheet = workbook[sheet_name]
    allocation = workbook["Allocation"]
    # Loop through each row in the worksheet
    # (values_only=True returns the cell values directly rather than Cell objects)
    id = 1
    tasks = {}
    excel_to_new_id = {}
    for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        # Process the row (here we simply print it)
        if row[35] != None and row[35] not in ["error", "Error", "Resizable", "Compact", "User Notes", "Dependency"]:
            excel_id = idx
            task_name = row[28].strip()
            dep, max_worker = extract_values(row[35])
            mh = row[38]
            task = {"id": id, "task_name": task_name, "base_effort": mh, "min": 1, "max": max_worker, "dependencies": dep, "resource": ""}
            tasks[excel_id] = task
            excel_to_new_id[excel_id] = id
            id += 1
    
    result = []
    for excel_id, task in tasks.items():
        # If the dependency is a list, convert each dependency.
        if isinstance(task["dependencies"], list):
            new_deps = [excel_to_new_id.get(dep, dep) for dep in task["dependencies"]]
        else:
            # Otherwise, treat it as a single dependency.
            new_deps = [excel_to_new_id.get(task["dependencies"], task["dependencies"])]
        task["dependencies"] = new_deps
        result.append(task)
    for task in result:
        for idx, row in enumerate(allocation.iter_rows(values_only=True), start=34):
            if row[13] is not None:
                if row[13].strip() == task["task_name"]:
                    for i in range(16, 100):
                        if row[i] != 0:
                            task["resource"] = allocation.cell(row=32, column=i).value
                            break
                    break
    with open("output.json", "w", encoding="utf-8") as f:
        json.dump(result, f, indent=4, ensure_ascii=False)
    
if __name__ == "__main__":
    
    jobs = {
    "SA/SD SolArch (JS8) - Solution Architect - DA": 1000,
    "SDU IntegEngin (JS6) - CBEV Solution Integrator": 1000,
    "SDU SolArch (JS6) - DevOps Tool Admin": 1000,
    "SDU SolArch (JS6) - Performance Lead": 1000,
    "SDU CPM (JS7) - DM CPM": 1000,
    "MA TestMgr (JS6) - Test lead": 1000,
    "SDU ITSysExp (JS5) - SW SME": 1000,
    "SDU IntegEngin (JS6) - CAF IE": 1000,
    "SDU SwDev (JS6) - CBEV Software Developer": 1000,
    "SA/SD TechSME (JS6) - NFR SA": 1000,
    "SDU SolArch (JS6) - Infra or SW Integration SA": 1000,
    "SDU SolArch (JS6) - DevOps SME": 1000,
    "SDU SolArch (JS6) - Security Master": 1000,
    "SA/SD SolArch (JS6) - CBEV Software Developer": 1000,
    "SDU CPM (JS7) - Rollout Manager": 1000,
    "SDU IntegEngin (JS4) - ST Test Engineer": 1000,
    "SDU SolArch (JS5) - RME or RTE": 1000,
    "SA/SD SolArch (JS7) - DM Lead / E2E SA": 1000,
    "SDU ITEngTest (JS4) - SIT Manual Tester": 1000,
    "SA/SD SolArch (JS7) - Performance SA": 1000,
    "MA TestMgr (JS6) - UAT Test Manager": 1000,
    "SDU ProgMgr (JS8) - Factory / Delivery Lead": 1000,
    "MA CPM (JS7) - PMO": 1000,
    "SDU ITEngTest (JS5) - ST Automation Test Engineer": 1000,
    "SDU ITSysExp (JS6) - LCM Support Team": 1000
    }
    {
    "SA/SD SolArch (JS8) - Solution Architect - DA": 118.97,
    "SDU IntegEngin (JS6) - CBEV Solution Integrator": 23.82,
    "SDU SolArch (JS6) - DevOps Tool Admin": 23.82,
    "SDU SolArch (JS6) - Performance Lead": 23.82,
    "SDU CPM (JS7) - DM CPM": 31.86,
    "MA TestMgr (JS6) - Test lead": 83.68,
    "SDU ITSysExp (JS5) - SW SME": 19.26,
    "SDU IntegEngin (JS6) - CAF IE": 23.82,
    "SDU SwDev (JS6) - CBEV Software Developer": 19.26,
    "SA/SD TechSME (JS6) - NFR SA": 92.78,
    "SDU SolArch (JS6) - Infra or SW Integration SA": 23.82,
    "SDU SolArch (JS6) - DevOps SME": 23.82,
    "SDU SolArch (JS6) - Security Master": 23.82,
    "SA/SD SolArch (JS6) - CBEV Software Developer": 37.1,
    "SDU CPM (JS7) - Rollout Manager": 31.86,
    "SDU IntegEngin (JS4) - ST Test Engineer": 14.66,
    "SDU SolArch (JS5) - RME or RTE": 23.82,
    "SA/SD SolArch (JS7) - DM Lead / E2E SA": 93.21,
    "SDU ITEngTest (JS4) - SIT Manual Tester": 14.66,
    "SA/SD SolArch (JS7) - Performance SA": 93.21,
    "MA TestMgr (JS6) - UAT Test Manager": 83.68,
    "SDU ProgMgr (JS8) - Factory / Delivery Lead": 44.43,
    "MA CPM (JS7) - PMO": 93.21,
    "SDU ITEngTest (JS5) - ST Automation Test Engineer": 14.66,
    "SDU ITSysExp (JS6) - LCM Support Team": 19.26
    }
    workbook = load_workbook(filename="Egon Thesis/real_project_data.xlsx", data_only=True)
    salaries = workbook["ExportCalc"]
    


